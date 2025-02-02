#from __future__ import unicode_literals
import re
import requests
import json
from django.conf import settings
import urllib
import uuid 
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, Http404,HttpResponseRedirect, JsonResponse
#from register.models import CustomUser, Team, Sport
from django.contrib.auth import authenticate, login, logout
from django.contrib import auth
from django.views.generic import View, ListView, FormView
from django.views import generic
from register.forms import LoginForm, TeamForm, PlayerForm
from django.db import IntegrityError
from django.db.models.functions import Lower
from django.contrib.auth.models import User
import json
from django.core import serializers, mail
from django.contrib.auth.decorators import login_required
from django.dispatch import receiver

from django.contrib.sites.shortcuts import get_current_site
from django.utils.encoding import force_bytes, force_text
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.template.loader import render_to_string

from register.tokens import account_activation_token
from django.core.mail import EmailMessage
from django.core import serializers

from random import choice
from string import ascii_uppercase

from io import StringIO
from xhtml2pdf import pisa
from django.template.loader import get_template
from django.template import Context
import re

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, Http404,HttpResponseRedirect, JsonResponse
from main.models import CustomUser, Team, Sport
from django.contrib.auth import authenticate, login, logout
from django.contrib import auth
from django.views.generic import View, ListView, FormView
from django.views import generic
from register.forms import LoginForm, TeamForm, PlayerForm
from django.db import IntegrityError
from django.contrib.auth.models import User
import json
from django.core import serializers, mail
from django.contrib.auth.decorators import login_required
from django.dispatch import receiver

from django.contrib.sites.shortcuts import get_current_site
from django.utils.encoding import force_bytes, force_text
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.template.loader import render_to_string

from register.tokens import account_activation_token
from django.core.mail import EmailMessage
from django.core import serializers

from random import choice
from string import ascii_uppercase
from io import StringIO
#import cStringIO as StringIO
from xhtml2pdf import pisa
from django.template.loader import get_template
from django.template import Context

from django.contrib.auth import get_user_model
User=get_user_model()

from django.conf import settings
from django.core.cache.backends.base import DEFAULT_TIMEOUT
from django.shortcuts import render
from django.views.decorators.cache import cache_page
from django.shortcuts import render
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status

from django.shortcuts import render
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from main.models import Group,Regplayer,Enteredplayer,Sport,Team, PaytmHistory, Amounts
from django.contrib.auth import logout
from django.contrib.auth.models import User
import json
from django.contrib.auth import get_user_model
import string
import random
from django.core import serializers
import pusher
from django.conf import settings
from django.core.cache.backends.base import DEFAULT_TIMEOUT
from django.shortcuts import render
from django.views.decorators.cache import cache_page, never_cache
from django.shortcuts import render
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponse
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from django.template.loader import get_template
from django.template import Context
from django.http import HttpResponse
from cgi import escape
#from weasyprint import HTML
from random import choice
from string import ascii_uppercase
import re
import openpyxl
from openpyxl.utils import get_column_letter
import csv
from django.utils.encoding import smart_str
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, Http404,HttpResponseRedirect, JsonResponse
from main.models import CustomUser, Team
from django.contrib.auth import authenticate, login, logout
from django.contrib import auth
from django.views.generic import View, ListView, FormView
from django.views import generic
from register.forms import LoginForm, TeamForm, PlayerForm
from django.db import IntegrityError
from django.contrib.auth.models import User
import json
from django.core import serializers, mail
from django.contrib.auth.decorators import login_required
from django.dispatch import receiver

from random import choice
from string import ascii_uppercase

from django.contrib.sites.shortcuts import get_current_site
from django.utils.encoding import force_bytes, force_text
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.template.loader import render_to_string
from register.tokens import account_activation_token
from django.core.mail import EmailMessage
from django.core import serializers
from main.models import Group,Regplayer,Enteredplayer,Sport,Money,Billcontrols,Pcradmin_user
# connection = mail.get_connection()
from django.contrib.auth import get_user_model
User=get_user_model()

pusher_client = pusher.Pusher(
  app_id='551467',
  key='a7ef26b891af0311773e',
  secret='9270c6a4ed1b9cfef8d0',
  cluster='ap2',
  ssl=True
)

CACHE_TTL = getattr(settings, 'CACHE_TTL', DEFAULT_TIMEOUT)

def is_pcradmin_admin(user):
	if user:
		if Pcradmin_user.objects.get(pk=1).user == user:
			return True
	return False

@cache_page(CACHE_TTL)
@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def index(request):
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	return render(request,'pcradmin/pcrhome1.html')


@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def render_pcrmail(request):
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		playerlist = User.objects.filter(grp_leader=1,is_active=True,deleted=0).order_by(Lower('name'))
		#playerlist=User.objects.all()
		data=[]
		for dt in playerlist:
			s=[]
			s.append(dt.name)
			s.append(dt.team.college)
			s.append(dt.team.city)
			s.append(dt.team.state)
			s.append(dt.email)
			s.append(dt.pk)
			s.append(dt.team.pk)
			if dt.team.activate==1:
				for i in range(40):
					if dt.team.confirmedsp1[i]>='1':
						data.append(s)
						break
		return JsonResponse({'groupleaders':data})

@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def modify_pcrmail(request):
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		resp={'success':1}
		data = json.loads( request.body.decode('utf-8') )
		mail_subject = data['sub']
		message = data['body']
		to_email = data['email_arr']
		for mail in to_email:
			email = EmailMessage(mail_subject, message, to=[mail])
			#try:
			email.send()
			#except:
			#	resp={'success':0}
		return JsonResponse(resp)

@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def send(request):
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		data=json.loads(request.body.decode('utf-8'))
		team=Team.objects.get(pk=data['clg_id'])
		#team=Team.objects.get(activate=1)
		splist=Sport.objects.all().order_by(Lower('sport'))
		players=User.objects.filter(team=team,deleted=0).order_by(Lower('name'))
		d=[]
		for sp in splist:
				s=[]
			#if team.confirmedsp1[sp.idno]>='1':
				s.append(sp.sport)
				s.append(sp.pk)
				p2=[]
				for pl in players:
					p=[]
					if pl.sportid[sp.idno]>='2':
						p.append(pl.name)
						p.append(pl.phone)
						p.append(pl.email)
						p.append(pl.pk)
						p2.append(p)
				s.append(p2)
				if len(p2)>0:
					d.append(s)
		return JsonResponse({'data':d})

@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def finalmail(request):
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		data=json.loads(request.body.decode('utf-8'))
		team=Team.objects.get(pk=data['clg_id'])
		ld=User.objects.get(team=team,grp_leader=1,deleted=0)
		ulist=User.objects.filter(team=team,deleted=0)
		success=1
		for sp in data['sport_id_arr']:
			sprt=Sport.objects.get(pk=sp)
			nm=[]
			mailid=[]
			for up in ulist:
				if up.sportid[sprt.idno]>='2':
					nm.append(up.name)
					mailid.append(up.email)
			message = render_to_string('pcradmin/msg1.html', {
													'college':team.college, 
													'sport':sprt.sport,
													'nmlist':nm,
													
													})
			mail_subject = 'BOSM 2018 | Confirmation Mail'
			email = EmailMessage(mail_subject, message, to=[ld.email])
			email.attach_file('Rate Sheet.pdf')
			email.content_subtype = "html"
			try:
				email.send()
			except:
				success=0


			try:
				cap=User.objects.get(team=team,captain=sp,deleted=0)
			except:
				#for toemail in mailid:
					message = render_to_string('pcradmin/msg1.html', {
															'college':team.college, 
															'sport':sprt.sport,
															'nmlist':nm,
															
															})
					mail_subject = 'BOSM 2018 | Confirmation Mail'
					email = EmailMessage(mail_subject, message, to=mailid)
					email.attach_file('Rate Sheet.pdf')
					email.content_subtype = "html"
					try:
						email.send()
					except:
						success=0
			else:
				message = render_to_string('pcradmin/msg1.html', {
														'college':team.college, 
														'sport':sprt.sport,
														'nmlist':nm,
														
														})
				mail_subject = 'BOSM 2018 | Confirmation Mail'
				email = EmailMessage(mail_subject, message, to=[cap.email])
				email.attach_file('Rate Sheet.pdf')
				email.content_subtype = "html"
				try:
					email.send()
				except:
					success=0
		resp={'success':success}
		return JsonResponse(resp)



 
#FOR ACTIVATION OF GRP LEADERS
@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def activateGrp(request):
	#receive pk of grp leader1
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		data = json.loads( request.body.decode('utf-8') )
		nm =data['pk']
		up=User.objects.get(pk=nm,grp_leader=1,deleted=0)
		#if data['task']=='activate':
		up.team.activate= 1
		list1=User.objects.filter(team=up.team,grp_leader=1,deleted=0)
		for i in list1:
				if i==up:
					pass
				else:
					i.deleted=1
					i.save()
		# else:
		# 	up.team.activate=0
		# 	list1=User.objects.filter(team=up.team,grp_leader=1,deleted=1)
		# 	for i in list1:
		# 		if i==up:
		# 			pass
		# 		else:
		# 			i.deleted=0
		# 			i.save()
		try:
			up.team.save()
			#up.save()
		except:
			resp={'success':0}
		else:
			resp={'success':1}
			message = render_to_string('pcradmin/msg2.html', {
															'username':up.username, 
															'user':up.name,
															
															})
			mail_subject = 'BOSM 2018 | Account activated'
			email = EmailMessage(mail_subject, message, to=[up.email])
			email.content_subtype = "html"
			email.send()
		stats_update_data = [7,3]
		pusher_client.trigger('my-channel7', 'my-event7', stats_update_data)
		return JsonResponse(resp)

@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def deactivateGrp(request):
	#receive pk of grp leader1
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		data = json.loads( request.body.decode('utf-8') )
		nm =data['pk']
		up=User.objects.get(pk=nm,grp_leader=1,deleted=0)
		# if data['task']=='activate':
		# 	up.team.activate= 1
		# 	list1=User.objects.filter(team=up.team,grp_leader=1,deleted=0)
		# 	for i in list1:
		# 		if i==up:
		# 			pass
		# 		else:
		# 			i.deleted=1
		# 			i.save()
		# else:
		up.team.activate=0
		list1=User.objects.filter(team=up.team,grp_leader=1,deleted=1)
		for i in list1:
				if i==up:
					pass
				else:
					i.deleted=0
					i.save()
		try:
				up.team.save()
		except:
			resp={'success':0}
		else:
			resp={'success':1}
		stats_update_data2 = [7,3]
		pusher_client.trigger('my-channel7', 'my-event7', stats_update_data2)
		return JsonResponse(resp)

@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def viewGrpLeaders(request):
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':

		list1=Team.objects.filter(activate=0).order_by(Lower('college'))
		list2=Team.objects.filter(activate=1).order_by(Lower('college'))
		d=[]
		for i in list1:
			grpld=User.objects.filter(grp_leader=1,deleted=0,team=i).order_by(Lower('name'))
			if grpld.count():#display only if someone has registered in the college
				s=[]
				s.append(i.college)
				s.append(i.city)
				s.append(i.state)
				s.append(i.pk)
				
				l=[]
				for j in grpld:
					l2=[]
					l2.append(j.pk)
					l2.append(j.name)
					l2.append(j.phone)
					l2.append(j.email)
					sprt=Sport.objects.all()
					sp=[]
					for k in sprt:
						if j.sportid[k.idno]>='1':
							sp.append(k.sport)
					l2.append(sp)
					if (j.coach):
							l2.append("YES")
					else:
							l2.append("NO")
					l2.append(j.gender)
					l.append(l2)
				s.append(l)
				d.append(s)

		d2=[]
		for i in list2:
				s=[]
				s.append(i.college)
				s.append(i.city)
				s.append(i.state)
				s.append(i.pk)
				grpld=User.objects.filter(grp_leader=1,team=i).order_by(Lower('name'))
				l=[]
				for j in grpld:
					l2=[]
					l2.append(j.pk)
					l2.append(j.name)
					l2.append(j.phone)
					l2.append(j.email)
					sprt=Sport.objects.all()
					sp=[]
					for k in sprt:
						if j.sportid[k.idno]>='1':
							sp.append(k.sport)
					l2.append(sp)
					if (j.coach):
						l2.append("YES")
					else:
						l2.append("NO")
					l2.append(j.gender)
					l2.append(j.deleted)
					l.append(l2)
				s.append(l)
				d2.append(s)
		return JsonResponse({'data':d,'data2':d2})




# FOR SPORTS LIMITS
@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def displaySp(request):
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		sports=Sport.objects.all().order_by(Lower('sport'))
		data=[]
		for sp in sports:
			s=[]
			s.append(sp.pk)
			s.append(sp.sport)
			s.append(sp.lower)
			s.append(sp.upper)
			
			data.append(s)
		#return HttpResponse(data1,content_type='application/json')
		return JsonResponse( {'limits':data})

@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def changeLimit(request):
	#get name of sport and both limits
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		data= json.loads(request.body.decode('utf-8'))
		sp=Sport.objects.get(pk=data['idno'])
		sp.lower=int(data['lowerLimit'])
		sp.upper=int(data['upperLimit'])
		try:
			sp.save()
		except:
			resp={'success':0}
		else:
			resp={'success':1}
		return JsonResponse(resp)

@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def addSp(request):
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		data= json.loads(request.body.decode('utf-8'))
		count=Sport.objects.all().count()
		sp=Sport.objects.create(idno=count+1,sport=data['sportName'],lower=data['lowerLimit'],upper=data['upperLimit'],gender=data['gender'])
		try:
			sp.save()
		except:
			resp={'success':0}
		else:
			resp={'success':1}
		return JsonResponse(resp)




# ADD college
@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def addTeam(request):
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		data= json.loads(request.body.decode('utf-8'))
		sp=Team.objects.create(college=data['clgName'],city=data['city'],state=data['state'])
		try:
			sp.save
		except:
			resp={'success':0}
		else:
			resp={'success':1}
		return JsonResponse(resp)




#VIEW STATS
@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def statssport(request):
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		d=json.loads(request.body.decode('utf-8'))
		sprt=Sport.objects.get(pk=d['idno'])
		tm=Team.objects.filter(activate=1).order_by(Lower('college'))
		data=[]
		mt=0
		ft=0
		ct=0
		mc=0
		fc=0
		cc=0
		tt=0
		tc=0
		for t in tm:
			up=User.objects.filter(team=t,deleted=0)
			print(up)
			maleTotal=0
			femaleTotal=0
			coachTotal=0
			maleconfirmed=0
			femaleconfirmed=0
			#maleconfirmed=up.filter(gender='male',coach=0,confirm1=1).count()
			#femaleconfirmed=up.filter(gender='female',coach=0,confirm1=1).count()
			coachconfirmed=0
			s=[]
			
			for u in up:
				if u.sportid[sprt.idno]>='1':
					if u.coach:
						coachTotal+=1
						if u.confirm1:
							coachconfirmed+=1
					else:
						if u.gender=='male':
							maleTotal+=1;
						elif u.gender=='female':
							femaleTotal+=1
						if u.sportid[sprt.idno]>='2':
							if u.gender=='male':
								maleconfirmed+=1;
							elif u.gender=='female':
								femaleconfirmed+=1

			mt+=maleTotal
			mc+=maleconfirmed
			ft+=femaleTotal
			fc+=femaleconfirmed
			ct+=coachTotal
			cc+=coachconfirmed
			if maleTotal+femaleTotal!=0 or coachTotal!=0:
				s.append(t.pk)
				s.append(t.college)
				s.append(t.city)
				s.append(t.state)
				s.append(maleconfirmed)
				s.append(maleTotal)
				s.append(femaleconfirmed)
				s.append(femaleTotal)
				s.append(maleconfirmed+femaleconfirmed)
				s.append(maleTotal+femaleTotal)
				s.append(coachconfirmed)
				s.append(coachTotal)
				
				
				
				
				
				data.append(s)

		total=[]
		
		total.append(mc)
		total.append(mt)
		total.append(fc)
		total.append(ft)
		total.append(mc+fc)
		total.append(mt+ft)
		total.append(cc)
		total.append(ct)
		
		
		resp={'college':data, 'total':total}
		return JsonResponse(resp)

@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def statscollege(request):
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		d=json.loads(request.body.decode('utf-8'))
		tm=Team.objects.get(pk=d['idno'])
		sp=Sport.objects.all().order_by(Lower('sport'))
		data=[]
		mt=0
		ft=0
		ct=0
		mc=0
		fc=0
		cc=0
		for t in sp:
			up=User.objects.filter(team=tm,deleted=0)
			maleTotal=0
			femaleTotal=0
			coachTotal=0
			maleconfirmed=0
			femaleconfirmed=0
			
			coachconfirmed=0
			s=[]

			mt=up.filter(gender='male',coach=0).count()
			mc=up.filter(gender='male',coach=0).exclude(confirm1=0).count()
			ft=up.filter(gender='female',coach=0).count()
			fc=up.filter(gender='female',coach=0).exclude(confirm1=0).count()
			tt=up.filter(coach=0).count()
			tc=up.filter(coach=0).exclude(confirm1=0).count()

			for u in up:
				if u.coach==t.pk:
					coachTotal+=1
					if u.confirm1:
						coachconfirmed+=1
				else:
					if u.sportid[t.idno]>='1':
						if u.gender=='male':
							maleTotal+=1;
						elif u.gender=='female':
							femaleTotal+=1
					if u.sportid[t.idno]>='2':
						if u.gender=='male':
							maleconfirmed+=1;
						elif u.gender=='female':
							femaleconfirmed+=1
			# mt+=maleTotal
			# mc+=maleconfirmed
			# ft+=femaleTotal
			# fc+=femaleconfirmed
			ct+=coachTotal
			cc+=coachconfirmed



			s.append(t.pk)
			s.append(t.sport)
			s.append(maleconfirmed)
			s.append(maleTotal)
			s.append(femaleconfirmed)
			s.append(femaleTotal)
			s.append(maleconfirmed+femaleconfirmed)
			s.append(maleTotal+femaleTotal)
			s.append(coachconfirmed)
			s.append(coachTotal)
			
			
			
			
			
			data.append(s)

		total=[]
		
		total.append(mc)
		total.append(mt)
		total.append(fc)
		total.append(ft)
		total.append(tc)
		total.append(tt)
		total.append(cc)
		total.append(ct)

		resp={'sport':data,'total':total}
		return JsonResponse(resp)

@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def stats(request):
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		tm=Team.objects.filter(activate=1).order_by(Lower('college'))
		total=[]
		data=[]
		mt=0
		ft=0
		ct=0
		mc=0
		fc=0
		cc=0
		for t in tm:
			up=User.objects.filter(team=t,deleted=0)
			maleTotal=up.filter(gender='male',coach=0).count()
			femaleTotal=up.filter(gender='female',coach=0).count()
			coachTotal=0
			maleconfirmed=up.filter(gender='male',coach=0).exclude(confirm1=0).count()
			femaleconfirmed=up.filter(gender='female',coach=0).exclude(confirm1=0).count()
			coachconfirmed=0


			s=[]
			
			for u in up:
				if u.coach:
					coachTotal+=1
					if u.confirm1:
						coachconfirmed+=1

			mt+=maleTotal
			mc+=maleconfirmed
			ft+=femaleTotal
			fc+=femaleconfirmed
			ct+=coachTotal
			cc+=coachconfirmed 

			s.append(t.pk)
			s.append(t.college)
			s.append(t.city)
			s.append(t.state)
			s.append(maleconfirmed)
			s.append(maleTotal)
			s.append(femaleconfirmed)
			s.append(femaleTotal)
			s.append(maleconfirmed+femaleconfirmed)
			s.append(maleTotal+femaleTotal)
			s.append(coachconfirmed)
			s.append(coachTotal)
			
			
			
			
			
			data.append(s)

			
			
		total.append(mc)
		total.append(mt)
		total.append(fc)
		total.append(ft)
		total.append(mc+fc)
		total.append(mt+ft)
		total.append(cc)
		total.append(ct)

		sp=Sport.objects.all().order_by(Lower('sport'))
		data2=[]
		for t in sp:
			up=User.objects.filter(deleted=0)
			

			maleTotal=0
			femaleTotal=0
			coachTotal=0
			maleconfirmed=0
			femaleconfirmed=0
			#maleconfirmed=up.filter(gender='male',coach=0,confirm1=1).count()
			#femaleconfirmed=up.filter(gender='female',coach=0,confirm1=1).count()
			coachconfirmed=0
			s=[]
			
			for u in up:
				if u:
					if u.sportid[t.idno]>='1':
						if u.team:
							if u.team.activate==1:
								if u.coach:
									coachTotal+=1
									if u.confirm1:
										coachconfirmed+=1
								else:
									if u.sportid[t.idno]>='2':
										if u.gender=='male':
											maleconfirmed+=1
											maleTotal+=1
										elif u.gender=='female':
											femaleconfirmed+=1
											femaleTotal+=1
									else:
										if u.gender=='male':
											maleTotal+=1
										elif u.gender=='female':
											femaleTotal+=1
			s.append(t.pk)
			s.append(t.sport)
			s.append(maleconfirmed)
			s.append(maleTotal)
			s.append(femaleconfirmed)
			s.append(femaleTotal)
			s.append(maleconfirmed+femaleconfirmed)
			s.append(maleTotal+femaleTotal)
			s.append(coachconfirmed)
			s.append(coachTotal)
			
			
			
			
			
			data2.append(s)
		resp={'college':data,'sport':data2,'total':total}

		return JsonResponse(resp)


@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def statscollegesport(request):
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		d=json.loads(request.body.decode('utf-8'))
		team=Team.objects.get(pk=d['col_id'])
		sport=Sport.objects.get(pk=d['sport_id'])
		user=User.objects.filter(team=team,deleted=0).order_by(Lower('name'))
		data=[]
		for u in user:
			p=[]
			if u.sportid[sport.idno]=='1':
				p.append(u.name)
				p.append(u.phone)
				p.append(u.email)
				p.append(u.gender)
				if u.captain==0 and u.coach==0:
					p.append("Participant")
				elif u.captain!=0 and u.coach==0:
					p.append("Captain")
				elif u.captain==0 and u.coach!=0:
					p.append("Coach")
				else:
					p.append("undefined")
				p.append(0)
				if u.pay2 or u.pay3:
					p.append(2)#payment status
				elif u.pay1:
					p.append(1)
				else:
					p.append(0)#payment status
				p.append(u.confirm1)
				data.append(p)


			if u.sportid[sport.idno]>='2':
				p.append(u.name)
				p.append(u.phone)
				p.append(u.email)
				p.append(u.gender)
				if u.captain==0 and u.coach==0:
					p.append("Participant")
				elif u.captain!=0 and u.coach==0:
					p.append("Captain")
				elif u.captain==0 and u.coach!=0:
					p.append("Coach")
				else:
					p.append("undefined")
				p.append(1)
				if u.pay2 or u.pay3:
					p.append(2)#payment status
				elif u.pay1:
					p.append(1)
				else:
					p.append(0)
				p.append(u.confirm1)
				data.append(p)
		resp={'participants':data}
		return JsonResponse(resp)




# EDIT DETAILS
@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def editDisplay(request):
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		tm=Team.objects.all().order_by(Lower('college'))
		slist=Sport.objects.all().order_by(Lower('sport'))
		data=[]
		for t in tm:
			s=[]
			try:
				up=User.objects.get(team=t, grp_leader=1, is_active=True,deleted=0)
			except:
				pass
			else:
				if up.team.activate:
					s.append(up.name)
					s.append(t.college)
					s.append(t.city)
					s.append(t.state)
					s.append(up.phone)
					s.append(up.email)
					s.append(up.pk)
					s.append(up.team.pk)
					s.append(up.gender)
					spid=[]
					for s2 in slist:
						if up.sportid[s2.idno]>='1':
							spid.append(s2.pk)
					s.append(spid)

					
					data.append(s)
		d=[]
		for sp in slist:
			d.append(sp.pk)
			d.append(sp.sport)
		resp={'groupleaders':data,'sports':d}
		return JsonResponse(resp)

@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def editDetails(request):
	#get pk of user
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		data=json.loads(request.body.decode('utf-8'))
		tm=Team.objects.get(pk=data['clg_id'])
		up=User.objects.filter(team=tm,deleted=0,grp_leader=0).order_by(Lower('name'))
		slist=Sport.objects.all().order_by(Lower('sport'))
		#sending only participants name and sports
		#(not captains name of the sports he is in as its of no use, 
		#participant can be more than 1 sport therefore showing more than 1 captain name 
		#does not make sense and too much work )
		data=[]
		for u in up:
			s=[]
			s.append(u.name)
			sp=[]
			spid=[]
			for s2 in slist:
				if u.sportid[s2.idno]>='1':
					sp.append(s2.sport)
					spid.append(s2.pk)
			s.append(sp)
			s.append(u.phone)
			s.append(u.email)
			s.append(u.pk)
			s.append(u.gender)
			s.append(spid)
			data.append(s)
		resp={'data':data}
		return JsonResponse(resp)

@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def editName(request):
	#get pk of user and new name
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		data=json.loads(request.body.decode('utf-8'))
		up=User.objects.get(pk=data['idno'])
		splist=Sport.objects.all()
		success=1
		for sp in splist:

			if str(sp.pk) in data['selectedSports']:
				#sp=Sport.objects.get(pk=int(dt))
				if up.team.confirmedsp1[sp.idno]>='1':
					up.sportid=replaceindex(up.sportid,sp.idno,'2')
				else:
					up.sportid=replaceindex(up.sportid,sp.idno,'1')
			else:
				up.sportid=replaceindex(up.sportid,sp.idno,'0')
				if up.captain==sp.pk:
					up.captain=0
				if up.coach==sp.pk:
					up.coach=0
		up.email=data['email']
		up.gender=data['gender']
		up.name=data['name']
		up.phone=data['phone']
		
		try:
			up.save()
		except:
			success=0

		return JsonResponse({'success':success})
	#END




#confirm teams
@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def confirmTeamDetails(request):#show all available grps
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		up=User.objects.filter(grp_leader=1,is_active=True,deleted=0).order_by(Lower('name'))
		activated=[]
		for u in up:
			if u.team.activate:
				s=[]
				s.append(u.name)
				s.append(u.team.college)
				s.append(u.team.city)
				s.append(u.team.state)
				s.append(u.phone)
				s.append(u.pk)
				s.append(u.team.pk)
				activated.append(s)
		resp=({'groupleaders':activated})
		return JsonResponse(resp)


@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def confirmTeamDisplay(request):#show details of grp
	#get pk of selected grp_leader
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		data=json.loads(request.body.decode('utf-8'))
		tm=Team.objects.get(pk=data['clg_id'])
		user = User.objects.filter(team=tm,deleted=0).order_by(Lower('name'))
		confirmedList=[]
		notconfirmedList=[]
		sports=Sport.objects.all().order_by(Lower('sport'))
		for sp in sports:
			s=[]
			upCount=0
			for u in user:
				if u.sportid[sp.idno]>='1':
					upCount+=1
			coachCount=user.filter(coach=sp.pk).count()
			ttCount= upCount - coachCount
			if ttCount>0:
				s.append(sp.pk)
				try:
					up=user.get(captain=sp.pk)
				except:
					s.append(' ')
				else:
					s.append(up.name)
				#s.append(tm.college)
				s.append(ttCount)
				s.append(sp.sport)
				s.append(tm.pk)
				#s.append(up.pk)
				#s.append(sp.idno)
				a= sp.idno
				#print(up.team.confirmedsp1)
				if tm.confirmedsp1[a]>='1' and upCount:
					confirmedList.append(s)
				elif upCount:
					notconfirmedList.append(s)
		resp={'confirmed':confirmedList,'unconfirmed':notconfirmedList}
		return JsonResponse(resp)

@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def confirmTeam(request):
	
	#get pk of grp leader selected and idno of sport confirmed
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		# connection.open()
		success=1
		data=json.loads(request.body.decode('utf-8'))
		tm=Team.objects.get(pk=data['clg_id'])
		user=User.objects.filter(team=tm,deleted=0)
		for dt in data['id_arr']:

			sp=Sport.objects.get(pk=dt)
			nm=[]
			emails=[]
			#up.team.confirmedsp1[sp.idno]=1
			#up.team.save()
			for u in user:
				#u.confirmed1[sp.idno]=1
				if u.sportid[sp.idno]=='1':
					u.sportid=replaceindex(u.sportid,sp.idno,'2')
					nm.append(u.name)

					if u.confirm1==0 and u.docs:#when u.docs is empty it will go to else
						u.confirm1=2
					elif u.confirm1==0:
						u.confirm1=1
					tm.confirmedsp1=replaceindex(tm.confirmedsp1,sp.idno,'1')
					try:	
						u.save()
						tm.save()
					except:
						success=0
					if success:
						emails.append(u.email)
						# message = render_to_string('pcradmin/msg5.html', {
						# 							'college':tm.college, 
						# 							'sport':sp.sport,
						# 							'nm':u.name,
													
						# 							})
						# mail_subject = 'Action Request | Document Upload | BOSM 2018'
				
						# email = EmailMessage(mail_subject, message, to=[u.email])
						# email.content_subtype="html"
						
						# try:
						# 	email.send()
						# except:
						# 	pass
						

				try:	
					u.save()
					tm.save()
				except:
					success=0


			ld=User.objects.get(team=tm,deleted=0,grp_leader=1)
			if success:
				message = render_to_string('pcradmin/msg4.html', {
												'college':tm.college, 
												'sport':sp.sport,
												'nmlist':nm,
												
												})
				mail_subject = 'Action Request | Document Upload | BOSM 2018'
		
				email = EmailMessage(mail_subject, message, to=[ld.email])
				email.content_subtype="html"
				
				try:
					email.send()
				except:
					pass

				email_message = render_to_string('pcradmin/msg5.html', {})
				email_subject = 'Action Request | Document Upload | BOSM 2018'
				
				mail = EmailMessage(email_subject, email_message, to=emails)
				mail.content_subtype="html"
				
				try:
					mail.send()
				except:
					pass
		# connection.close()
				
		update_data3 = [9,2]
		pusher_client.trigger('dashboard-update', 'dashboard-update-event', update_data3)
		return JsonResponse({'success':success})

@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def unconfirmTeam(request):
	#get pk of grp leader selected and idno of sport confirmed
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		success=1
		data=json.loads(request.body.decode('utf-8'))
		tm=Team.objects.get(pk=data['clg_id'])
		user=User.objects.filter(team=tm,deleted=0)

		for dt in data['id_arr']:

			sp=Sport.objects.get(pk=dt)
			
			for u in user:
				if u.sportid[sp.idno]>='2':
					u.sportid=replaceindex(u.sportid,sp.idno,'1')
					c=0
					for i in range(40):
						if u.sportid[i]<'2':
							c+=1
						if c==40:
							u.confirm1=0
					tm.confirmedsp1=replaceindex(tm.confirmedsp1,sp.idno,'0')
				try:	
					u.save()
					tm.save()
				except:
					success=0
		update_data3 = [9,2]
		pusher_client.trigger('dashboard-update', 'dashboard-update-event', update_data3)
		return JsonResponse({'success':success})




#EXCEL1
@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def excelDisplay(request):
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		up=User.objects.filter(grp_leader=1,deleted=0).order_by(Lower('name'))
		data=[]
		for u in up:
			s=[]
			s.append(u.name)
			s.append(u.team.college)
			s.append(u.team.city)
			s.append(u.team.state)
			s.append(u.pk)
			data.append(s)
		return JsonResponse({'groupleaders':data})

@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def createExcel(request,pk):
	#get pk of grp l
	#data=json.loads(request.body.decode('utf-8'))
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	import openpyxl
	from openpyxl.utils import get_column_letter
	leader = User.objects.get(pk=pk)
	#leader=User.objects.get(grp_leader=1)
	queryset=User.objects.filter(team=leader.team,deleted=0).order_by(Lower('name'))#select queryset
	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = 'attachment; filename=database.xlsx'
	wb = openpyxl.Workbook()
	ws = wb.get_active_sheet()
	ws.title = leader.team.college

	row_num = 0

	columns = [
		(u"Name", 30),
		(u"Email", 20),
		(u"Phone", 15),
		(u"Sports", 30),
		(u"Captain", 15),
		(u"Coach", 15),
	]

	for col_num in range(len(columns)):
		c = ws.cell(row=row_num + 1, column=col_num + 1)
		c.value = columns[col_num][0]
		#c.style.font.bold = True
		# set column width
		ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]
	spt=Sport.objects.all()
	for obj in queryset:
		row_num += 1
		captn=''
		if obj.captain:
			c1=Sport.objects.get(idno=obj.captain)
			captn=c1.sport
		cch=''
		if obj.coach:
			c2=Sport.objects.get(idno=obj.coach)
			cch=c2.sport
		sprt=''
		for i in spt:
			if obj.sportid[i.idno]>='1':
				sprt=sprt+i.sport
		row = [
			obj.name,
			obj.email,
			obj.phone,
			sprt,
			captn,
			cch,

		]

		for col_num in range(len(row)):
			c = ws.cell(row=row_num + 1, column=col_num + 1)
			c.value = row[col_num]
			#c.style.alignment.wrap_text = True

	wb.save(response)
	return response

@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def leaderExcel(request):
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	#get pk of grp l
	import openpyxl
	from openpyxl.utils import get_column_letter
	#leader = User.objects.get(pk=pk)
	#leader=User.objects.get(grp_leader=1)
	queryset=User.objects.filter(grp_leader=1,deleted=0).order_by(Lower('name'))#select queryset
	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = 'attachment; filename=leaderdatabase.xlsx'
	wb = openpyxl.Workbook()
	ws = wb.get_active_sheet()
	ws.title = "groupleader sheet"

	row_num = 0

	columns = [
		(u"ID", 15),
		(u"Name", 40),
		(u"Email", 50),
		(u"Phone", 20),
		(u"College", 20),
		(u"No of teams", 20),#no of teams in which confirmed?
		(u"No of players", 10),
	]

	for col_num in range(len(columns)):
		c = ws.cell(row=row_num + 1, column=col_num + 1)
		c.value = columns[col_num][0]
		#c.style.font.bold = True
		# set column width
		ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]
	
	for obj in queryset:
		#if obj.captain:
			#sp=Sport.objects.get(idno=obj.captain)
			#count=queryset.filter(sport=sp).count()
			count=User.objects.filter(team=obj.team, confirm1=1,coach=0,deleted=0).count()
			sportcount=0
			for c in obj.team.confirmedsp1:
				if c>='1':
					sportcount+=1
			row_num += 1
			row = [
				obj.pk,
				obj.name,
				obj.email,
				obj.phone,
				obj.team.college,
				sportcount,
				count,
				

			]

			for col_num in range(len(row)):
				c = ws.cell(row=row_num + 1, column=col_num + 1)
				c.value = row[col_num]
				#c.style.alignment.wrap_text = True

	wb.save(response)
	return response

@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def leaderCsv(request):
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	import csv
	from django.utils.encoding import smart_str
	queryset=User.objects.filter(grp_leader=1,deleted=0).order_by(Lower('name'))#select queryset
	response = HttpResponse(content_type='text/csv')
	response['Content-Disposition'] = 'attachment; filename=groupleader.csv'
	writer = csv.writer(response, csv.excel)
	response.write(u'\ufeff'.encode('utf8')) # BOM (optional...Excel needs it to open UTF-8 file properly)
	writer.writerow([
		smart_str(u"ID"),
		smart_str(u"Name"),
		smart_str(u"Email"),
		smart_str(u"Phone"),
		smart_str(u"Collge"),
		smart_str(u"No of teams"),
		smart_str(u"No of players"),
	])
	for obj in queryset:
		count=User.objects.filter(team=obj.team, confirm1=1,coach=0,deleted=0).count()
		sportcount=0
		for c in obj.team.confirmedsp1:
				if c>='1':
					sportcount+=1
		writer.writerow([
		smart_str(obj.pk),
		smart_str(obj.name),
		smart_str(obj.email),
		smart_str(obj.phone),
		smart_str(obj.team.college),
		sportcount,
		count,
		])
	return response

@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def createCsv(request,pk):
	#data=json.loads(request.body.decode('utf-8'))
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	import csv
	from django.utils.encoding import smart_str
	leader = User.objects.get(pk=pk)
	queryset=User.objects.filter(team=leader.team,deleted=0).order_by(Lower('name'))
	response = HttpResponse(content_type='text/csv')
	response['Content-Disposition'] = 'attachment; filename=mymodel.csv'
	writer = csv.writer(response, csv.excel)
	response.write(u'\ufeff'.encode('utf8')) # BOM (optional...Excel needs it to open UTF-8 file properly)
	writer.writerow([
		smart_str(u"Name"),
		smart_str(u"Email"),
		smart_str(u"Phone"),
		smart_str(u"Sports"),
		smart_str(u"Captain"),
		smart_str(u"Coach"),
	])

	spt=Sport.objects.all()
	for obj in queryset:
		captn=''
		if obj.captain:
			c1=Sport.objects.get(idno=obj.captain)
			captn=c1.sport
		cch=''
		if obj.coach:
			c2=Sport.objects.get(idno=obj.coach)
			cch=c2.sport
		sprt=''
		for i in spt:
			if obj.sportid[i.idno]>='1':
				sprt=sprt+i.sport
		writer.writerow([
		smart_str(obj.name),
		smart_str(obj.email),
		smart_str(obj.phone),
		smart_str(sprt),
		smart_str(captn),
		smart_str(cch),
		])

	return response

@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def leaderPdf(request):
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	data = []
	queryset=User.objects.filter(grp_leader=1,deleted=0).order_by(Lower('name'))
	for obj in queryset:
			count=User.objects.filter(team=obj.team, confirm1=1,coach=0,deleted=0).count()
			sportcount=0
			for c in obj.team.confirmedsp1:
				if c>='1':
					sportcount+=1
			data.append({"pk":obj.pk,"name":obj.name,"college":obj.team.college,"mobile_no":obj.phone,"email_id":obj.email,"sportcount":sportcount,"players":count})
	context = {"leaderlist":data}
	return render(request,'pcradmin/leaderpdfstats.html',context)

@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def createPdf(request,pk):
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	data = []
	leader = User.objects.get(pk=pk)
	queryset=User.objects.filter(team=leader.team,deleted=0).order_by(Lower('name'))
	spt=Sport.objects.all()
	for obj in queryset:
		captn=''
		if obj.captain:
			c1=Sport.objects.get(idno=obj.captain)
			captn=c1.sport
		cch=''
		if obj.coach:
			c2=Sport.objects.get(idno=obj.coach)
			cch=c2.sport
		sprt=''
		for i in spt:
			if obj.sportid[i.idno]>='1':
				sprt=sprt+i.sport+" "
		data.append({"name":obj.name,"mobile_no":obj.phone,"email_id":obj.email,"sport":sprt,"Captain":captn,"Coach":cch})
	context = {"teamlist":data}
	return render(request,'pcradmin/collegepdfstats.html',context)


#RESEND CREDENTIALS
@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def credleader(request):
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		tmlist=Team.objects.filter(activate=1).order_by(Lower('college'))
		d=[]
		for t in tmlist:
			try:
				gl=User.objects.get(grp_leader=1,team=t,deleted=0)
			except:
				pass
			else:
				s=[]
				s.append(gl.name)
				s.append(gl.team.college)
				s.append(gl.team.city)
				s.append(gl.team.state)
				s.append(gl.email)
				s.append(gl.pk)
				s.append(gl.team.pk)
				
				d.append(s)
		return JsonResponse({'groupleaders': d})

@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def credplayers(request):
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		data=json.loads(request.body.decode('utf-8'))
		t=Team.objects.get(pk=data['clg_id'])
		splist=Sport.objects.all().order_by(Lower('sport'))
		players=User.objects.filter(team=t,deleted=0).order_by(Lower('name'))
		d=[]
		for sp in splist:
				s=[]
			#if team.confirmedsp1[sp.idno]>='1':
				s.append(sp.sport)
				s.append(sp.pk)
				p2=[]
				for pl in players:
					p=[]
					if pl.sportid[sp.idno]>='1':
						p.append(pl.name)
						p.append(pl.phone)
						p.append(pl.email)
						p.append(pl.pk)
						p2.append(p)
				s.append(p2)
				d.append(s)
		return JsonResponse({'data':d})

@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def sendcred(request):
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		data = json.loads( request.body.decode('utf-8') )
		resp={'success':1}
		for i in data['id_arr']:
			up=User.objects.get(pk=i)
			up.username= (''.join(choice(ascii_uppercase) for i in range(5))) + str(up.pk)
			passworduser=(''.join(choice(ascii_uppercase) for i in range(12)))+str(up.pk)
			up.set_password(passworduser)
			try:
				up.save()
				#print('Hellos')
			except:
				resp={'success':0}
			else:
				#print(up.username)
				message = render_to_string('pcradmin/msg3.html', {
														'user':up.name, 
														'username':up.username,
														'password':passworduser,
														
														})
				mail_subject = 'Your account details for BOSM \'18'
				email = EmailMessage(mail_subject, message, to=[up.email])
				email.content_subtype = "html"
				try:
					email.send()
				except:
					resp={'success':0}
		return JsonResponse(resp)




#CHANGE GROUP LEADER 
@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def changeleader(request):
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		data=json.loads(request.body.decode('utf-8'))
		tm=Team.objects.get(pk=data['clg_id'])
		uplist=User.objects.filter(team=tm,deleted=0,grp_leader=0).order_by(Lower('name'))
		splist=Sport.objects.all().order_by(Lower('sport'))
		d2=[]
		for u in uplist:
			s=[]
			s.append(u.pk)
			s.append(u.name)
			s.append(u.phone)
			s.append(u.email)
			d=[]
			for sp in splist:
				if u.sportid[sp.idno]>='1':
					d.append(sp.sport)
			s.append(d)
			s.append(u.gender)
			if u.captain==0 and u.coach==0:
				s.append("Participant")
			elif u.captain!=0 and u.coach==0:
				s.append("Captain")
			elif u.captain==0 and u.coach!=0:
				s.append("Coach")
			else :
				s.append=("")
			d2.append(s)
		return JsonResponse({'data':d2})

@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def changegl(request):
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		data=json.loads(request.body.decode('utf-8'))
		old=User.objects.get(pk=data['old_id'])
		new=User.objects.get(pk=data['new_id'])
		new.grp_leader=1
		old.grp_leader=0
		success=1
		try:
			new.save()
			old.save()
		except:
			success=0
		#print(success)
		return JsonResponse({'success':success})




#DOCUMENT VERIFICATION
#for grpleader use render_pcrmail
@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def docurl(request):
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		data=json.loads(request.body.decode('utf-8'))
		tm=Team.objects.filter(pk=data['clg_id'])
		uplist=User.objects.filter(team=tm, deleted=0).order_by(Lower('name'))
		splist=Sport.objects.all().order_by(Lower('sport'))
		d2=[]
		d3=[]
		for u in uplist:
			s=[]
			d=[]
			if u.confirm1==1:
				s.append(u.pk)
				s.append(u.name)
				for sp in splist:
					if u.sportid[sp.idno]>='2':
						d.append(sp.sport)
				s.append(d)
				s.append("")
				s.append(0)
				s.append("")
				s.append(0)
				d2.append(s)
			if u.confirm1==2:
				s.append(u.pk)
				s.append(u.name)
				for sp in splist:
					if u.sportid[sp.idno]>='2':
						d.append(sp.sport)
				s.append(d)
				if bool(u.docs):
					s.append(u.docs.url)
					s.append(1)
				else:
					s.append("")
					s.append(0)
				if bool(u.docs2):
					s.append(u.docs2.url)
					s.append(1)
				else:
					s.append("")
					s.append(0)
				d2.append(s)
			if u.confirm1>=3:
				s.append(u.pk)
				s.append(u.name)
				for sp in splist:
					if u.sportid[sp.idno]>='2':
						d.append(sp.sport)
				s.append(d)
				if bool(u.docs):
					s.append(u.docs.url)
					#s.append(1)
				else:
					s.append("")
					#s.append(0)
				if bool(u.docs2):
					s.append(u.docs2.url)
					#s.append(1)
				else:
					s.append("")
					#s.append(0)
				d3.append(s)
		return JsonResponse({'unconfirmed':d2,'confirmed':d3})

@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def docapprove(request):
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		data=json.loads(request.body.decode('utf-8'))
		success=1
		tm=User.objects.get(pk=data['id_arr'][0]).team
		glmail=User.objects.get(team=tm,grp_leader=1,deleted=0).email
		nmlist=[]
		maillist=[]
		for ppk in data['id_arr']:
			u=User.objects.get(pk=ppk)
			u.confirm1=3
			u.save()
			nmlist.append(u.name)
			maillist.append(u.email)
			rp = Regplayer()
			rp.name = User.objects.get(pk=u.pk)
			rp.gender = u.gender
			rp.college = u.team.college
			rp.city = u.team.city
			rp.mobile_no = u.phone
			rp.email_id = u.email
			rp.sport=''
			for s in Sport.objects.all():
				if u.sportid[s.idno]=='2':
					rp.sport=rp.sport+s.sport+','
			rp.sport = rp.sport[:-1]
			rp.save()
			rp.uid = "18CB"+str(100000+rp.pk)[-4:]
			rp.save()
			
		message = render_to_string('pcradmin/msg7.html', {
														'college':tm.college, 
														
														})
		mail_subject = 'Documents Verified for BOSM \'18'
		email = EmailMessage(mail_subject, message, to=maillist)
		email.content_subtype = "html"
		# try:
		# 	#pass
		# 	email.send()
		# except:
		# 	pass
		message = render_to_string('pcradmin/msg8.html', {
														'nmlist':nmlist,
														'college':tm.college, 
														
														})
		mail_subject = 'Documents Verified for BOSM \'18'
		email = EmailMessage(mail_subject, message, to=[glmail])
		email.content_subtype = "html"
		# try:
		# 	email.send()
		# except:
		# 	pass
		return JsonResponse({'success':1})





#DASHBOARD
@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def dashboard(request):
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		upmale=User.objects.filter(deleted=0,gender='male').exclude(team__isnull=True)
		upfemale=User.objects.filter(deleted=0,gender='female').exclude(team__isnull=True)
		d1=[]
		d2=[]
		d3=[]
		d4=[]
		x=[]
		y=[]
		a=upmale.filter(confirm1=4,pay1=1).count()
		b=upfemale.filter(confirm1=4,pay1=1).count()
		x.append(a+b)
		x.append(a)
		x.append(b)
		x.append(a+b)
		d4.append(x)
		a=upmale.filter(confirm1=4,pay2=1).count()
		b=upfemale.filter(confirm1=4,pay2=1).count()
		a2=upmale.filter(confirm1=4,pay3=1).count()
		b2=upfemale.filter(confirm1=4,pay3=1).count()
		y.append(a+b+a2+b2)
		y.append(a+a2)
		y.append(b+b2)
		y.append(a+b+a2+b2)
		d4.append(y)
		d1.append(upmale.count())
		d1.append(upfemale.count())
		d1.append(upmale.count()+upfemale.count())
		maledoc1=upmale.filter(confirm1=2).count()
		femaledoc1=upfemale.filter(confirm1=2).count()
		maledoc2=upmale.filter(confirm1=3).count()
		femaledoc2=upfemale.filter(confirm1=3).count()
		maleconfirm=upmale.filter(confirm1=1).count()
		femaleconfirm=upfemale.filter(confirm1=1).count()
		d2.append(maledoc1+maleconfirm+maledoc2)
		d2.append(femaledoc1+femaleconfirm+femaledoc2)
		d2.append(maledoc1+maleconfirm+femaledoc1+femaleconfirm+maledoc2+femaledoc2)
		d3.append(maledoc1+maledoc2)
		d3.append(femaledoc1+femaledoc2)
		d3.append(maledoc1+femaledoc1+maledoc2+femaledoc2)
		return JsonResponse({'data1':d1,'data2':d2,'data3':d3,'data4':d4})




# delete users
@login_required(login_url='/regsoft/')
@user_passes_test(is_pcradmin_admin, login_url='/regsoft/')
def delete(request):
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method == 'POST':
		data=json.loads(request.body.decode('utf-8'))
		up = User.objects.get(pk=data['idno'])
		success=1
		try:
			up.delete()
		except:
			success=0
		return JsonResponse({'success': success})




# logout
def logoutView(request):
	if request.user.is_authenticated():
		pass
	else:
		return HttpResponseRedirect('/pcradmin/')
	logout(request)
	return HttpResponseRedirect('/')




#utilities
def replaceindex(text,index=0,replacement=''):
	return '%s%s%s'%(text[:index],replacement,text[index+1:])


@never_cache
def refresh(request):
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		p=PaytmHistory.objects.get(pk=request.POST['id'],STATUS='PENDING')
		prereg=Amounts.objects.get(name='pre').amount
		reg=Amounts.objects.get(name='reg').amount
		#p=PaytmHistory.objects.get(pk=3)
		URL = "https://pguat.paytm.com/oltp/HANDLER_INTERNAL/TXNSTATUS"
		URL += '?JsonData={"MID":"%s","ORDERID":"%s"}'%(p.MID,p.ORDERID)
		print(requests.get(URL),requests.get(URL).text)
		response_data = json.loads(requests.get(URL).text)
		if (response_data['STATUS']=='TXN_SUCCESS'):
			p.STATUS=response_data['STATUS']
			p.BANKTXNID=response_data['BANKTXNID']
			p.RESPCODE=response_data['RESPCODE']
			p.RESPMSG=response_data['RESPMSG']
			p.RESPMSG=response_data['RESPMSG']
			p.BANKNAME=response_data['BANKNAME']
			p.PAYMENTMODE=response_data['PAYMENTMODE']
			p.save()
			order_id=response_data['ORDERID']
			u1=User.objects.filter(orderid1=order_id)
			u2=User.objects.filter(orderid2=order_id)
			u3=User.objects.filter(orderid3=order_id)
			upre=[]
			ureg=[]
			up2r=[]
			premail=[]
			regmail=[]
			p2rmail=[]
			for u in u1:
				upre.append(u.name)
				premail.append(u.email)
				u.pay1=1
				u.pcramt=prereg
				u.confirm1=4
				u.save()
			for u in u2:
				ureg.append(u.name)
				regmail.append(u.email)
				u.pay2=1
				u.pcramt=reg
				u.confirm1=4
				u.save()
			for u in u3:
				up2r.append(u.name)
				p2rmail.append(u.email)
				u.pay3=1
				u.pcramt=reg
				u.confirm1=4
				u.save()
			gl=User.objects.get(pk=p.user)
			message = render_to_string('register/msg6.html', {
											'user':gl.name, 
											'prereg':upre,
											'reg':ureg,
											'prereg2reg':up2r,
											'college':gl.team.college,
											'amount':p.TXNAMOUNT,
											'TXNID':p.TXNID,
											'timestamp':p.TXNDATE,
											'status':p.STATUS,
											'orderid':order_id,
											
											})
			mail_subject = 'Your Payment Details | BOSM 2018'
			email = EmailMessage(mail_subject, message, to=["bosmpayments@gmail.com"]+[request.user.email]+premail+regmail+p2rmail)
			email.content_subtype = "html"
			email.send() 
		if (response_data['STATUS']=='TXN_FAILURE'):
			p.STATUS=response_data['STATUS']
			p.BANKTXNID=response_data['BANKTXNID']
			p.RESPCODE=response_data['RESPCODE']
			p.RESPMSG=response_data['RESPMSG']
			p.RESPMSG=response_data['RESPMSG']
			p.BANKNAME=response_data['BANKNAME']
			p.PAYMENTMODE=response_data['PAYMENTMODE']
			p.save()
			order_id=response_data['ORDERID']
			u1=User.objects.filter(orderid1=order_id)
			u2=User.objects.filter(orderid2=order_id)
			u3=User.objects.filter(orderid3=order_id)
			upre=[]
			ureg=[]
			up2r=[]
			premail=[]
			regmail=[]
			p2rmail=[]
			for u in u1:
				upre.append(u.name)
				premail.append(u.email)
			for u in u2:
				ureg.append(u.name)
				regmail.append(u.email)
			for u in u3:
				up2r.append(u.name)
				p2rmail.append(u.email)
			gl=User.objects.get(pk=p.user)
			message = render_to_string('register/msg6.html', {
											'user':gl.name, 
											'prereg':upre,
											'reg':ureg,
											'prereg2reg':up2r,
											'college':gl.team.college,
											'amount':p.TXNAMOUNT,
											'TXNID':p.TXNID,
											'timestamp':p.TXNDATE,
											'status':p.STATUS,
											'orderid':order_id,
											
											})
			mail_subject = 'Your Payment Details | BOSM 2018'
			email = EmailMessage(mail_subject, message, to=["bosmpayments@gmail.com"]+[request.user.email]+premail+regmail+p2rmail)
			email.content_subtype = "html"
			email.send() 
	
		return HttpResponseRedirect('/pcradmin/paymentdetails/')

@never_cache
def paydetails(request):
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	plist1=PaytmHistory.objects.filter(STATUS='PENDING')
	plist2=PaytmHistory.objects.filter(STATUS='TXN_SUCCESS')
	plist3=PaytmHistory.objects.filter(STATUS='TXN_FAILURE')
	d=[]
	for i in plist1:
		s=[]
		u=User.objects.get(pk=i.user)
		s.append(i.pk)
		s.append(i.TXNID)
		s.append(i.TXNAMOUNT)
		s.append(u.name)
		d.append(s)
		s.append(i.STATUS)
	for i in plist2:
		s=[]
		u=User.objects.get(pk=i.user)
		s.append(i.pk)
		s.append(i.TXNID)
		s.append(i.TXNAMOUNT)
		s.append(u.name)
		d.append(s)
		s.append(i.STATUS)
	for i in plist3:
		s=[]
		u=User.objects.get(pk=i.user)
		s.append(i.pk)
		s.append(i.TXNID)
		s.append(i.TXNAMOUNT)
		s.append(u.name)
		d.append(s)
		s.append(i.STATUS)
	return render(request,'pcradmin/paydetails.html',{'data':d})

def teammail(request):
	if request.user.is_authenticated():
		if is_pcradmin_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		data=json.loads(request.body.decode('utf-8'))
		tm=Team.objects.get(pk=data['id'])
		gl=User.objects.get(team=tm,grp_leader=1,deleted=0)
		ulist=User.objects.filter(team=tm,deleted=0).order_by(Lower('name'))
		nmlist=[]
		for i in ulist:
			if i.confirm1>=1:
				nmlist.append(i.name)
		message = render_to_string('pcradmin/msg9.html', {
								'college':tm.college, 
								'nmlist':nmlist, 
								})
		mail_subject = 'Final Confirmation for BOSM, 2018'
		email = EmailMessage(mail_subject, message, to=[gl.email])
		email.attach_file('Travel Rate List BOSM 2018.pdf')
		email.attach_file('Checklist.pdf')
		email.content_subtype = "html"
		success=0
		try:
			email.send() 
		except:
			success=1
		return JsonResponse({'success':success})
	
